import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import time
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import clean
companies = []
open_workbook = '/Users/umang/umang/3-1/random/stock_market/Top 500 Companies as on 31 March 2022 based on market capitalisation.xlsx'
workbook = openpyxl.load_workbook(open_workbook)
ws = workbook.active

f  = open("begin.txt", "w")
f.write('0')
#print(clean.remaining)
for rows in range(3,1126):
    if(ws.cell(row = rows, column = 2).value!=None):
        companies.append(ws.cell(row = rows, column = 2).value)
max_retries = 10  # Number of times to retry
a = 0  # Initialize your counter

#PATH = "/Users/umang/umang/3-1/random/chromedriver-mac-arm64/chromedriver"

# Create the webdriver instance with options

# Create the webdriver instance with options()\
driver = webdriver.Chrome()
def ball(begin):
    
    try:
        for i in range(begin,len(companies)):
        
            driver = webdriver.Chrome()

            driver.get("https://www.bseindia.com/markets/equity/EQReports/StockPrcHistori.aspx?flag=0&type=ETF")
            wait = WebDriverWait(driver, 20)
            search= wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_smartSearch_mf")))

            options = webdriver.ChromeOptions()



            search_radio = driver.find_element(By.ID, "ContentPlaceHolder1_rad_no1")
            search_radio.click()
            search = driver.find_element(By.ID, "ContentPlaceHolder1_smartSearch")
            search.click()
            search.send_keys(companies[i])
            search.send_keys(Keys.ENTER)
            search_dates = driver.find_element(By.ID, "ContentPlaceHolder1_txtFromDate")
            search_dates.click()
            search = driver.find_element(By.CLASS_NAME, "ui-datepicker-month")
            search.click()
            drop=Select(search)
            drop.select_by_visible_text("Jan")
            search = driver.find_element(By.CLASS_NAME, "ui-datepicker-year")
            search.click()
            drop=Select(search)
            drop.select_by_visible_text("2021")
            search = driver.find_element(By.CLASS_NAME, "ui-datepicker-calendar")
            date = search.find_element(By.XPATH, ".//a[text()='1']" )
            date.click()
            search_dates = driver.find_element(By.ID, "ContentPlaceHolder1_txtToDate")
            search_dates.click()
            search = driver.find_element(By.CLASS_NAME, "ui-datepicker-month")
            search.click()
            drop=Select(search)
            drop.select_by_visible_text("Nov")
            search = driver.find_element(By.CLASS_NAME, "ui-datepicker-year")
            search.click()
            drop=Select(search)
            drop.select_by_visible_text("2023")
            search = driver.find_element(By.CLASS_NAME, "ui-datepicker-calendar")
            date = search.find_element(By.XPATH, ".//a[text()='6']" )
            date.click()
            search = driver.find_element(By.ID, "ContentPlaceHolder1_btnSubmit")
            search.click()
            wait = WebDriverWait(driver, 20)
            search= wait.until(EC.visibility_of_element_located((By.ID, "ContentPlaceHolder1_btnDownload1")))

            #search = driver.find_element(By.ID, "ContentPlaceHolder1_btnDownload1")

            driver.execute_script("arguments[0].click();", search)
            f = open("begin.txt", "w")
            f.write(str(i))


            #search.click()

            #search.click()



            #search_radio.send_keys(Keys.ENTER)


            time.sleep(3)

            # search.send_keys("infosys ltd")
    except Exception as e:
        driver.refresh()
        a =0
        
            # Handle other exceptions if needed
        print(f"An error occurred on attempt {a + 1}: {e}")
        a += 1  # Increment the counter
while True:
    f = open("begin.txt", 'r')
    begin = f.read()
    begin = int(begin)
    driver.refresh()
    
    ball(begin)
