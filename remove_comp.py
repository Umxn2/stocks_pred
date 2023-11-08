import openpyxl
import os
open_workbook = 'eligible.xlsx'
workbook = openpyxl.load_workbook(open_workbook)
ws = workbook.active
count = 0
for i in range(3,2000):
    if(ws.cell(row = i, column = 2).value!=None):
        if(int(ws.cell(row = i, column = 2).value)>=543980):
            print(ws.cell(row = i, column = 3).value)
            count =count +1
            ws.delete_rows(idx = i)
            print(count)
            workbook.save(open_workbook)
