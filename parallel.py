import multiprocessing
import model_
import openpyxl
import joblib
from multiprocessing import Array
from multiprocessing import Pool, cpu_count
import ctypes
#print(model_.stocks(500325).final_score)
open_workbook = '/Users/umang/umang/3-1/random/stock_market/Top 500 Companies as on 31 March 2022 based on market capitalisation.xlsx'
def check_empty_row():
    i = 3
    while (ws.cell(row = i, column = 2).value!=None):
        i=i+1
    return i
workbook = openpyxl.load_workbook(open_workbook)
ws = workbook.active
companies = []
items = []
i = check_empty_row()
for rows in range(3,i):
    items.append(ws.cell(row = rows, column = 2).value)
#model_.stocks(items)

def append(b):
        companies.append(b)
        print("companies appended")
        return companies
    
def initializer_func(A):
    global a
    a = A
n = len(items)
A = Array(ctypes.py_object, n)
#print(items)
# for i in range(len(items)):
#      append(model_.stocks(items[i]))
#model_.stocks(items[i])
if __name__ == "__main__":
    # pool_obj = multiprocessing.Pool()
    # ans = pool_obj.map(append, (model_.stocks(items[b]) for b in range(len(items))))
    with Pool(processes=cpu_count(), initializer=initializer_func, initargs=(A,)) as pool:
        pool.map(append,( model_.stocks(items[b]) for b in range(len(items))))
    result =0
    print(f'parent: {list(A)}')
#append()
# i watch you while you sllep