import os
import openpyxl
file_path = "data_more"
present = []
open_workbook = '/Users/umang/Downloads/eligible (1).xlsx'
workbook = openpyxl.load_workbook(open_workbook)
ws = workbook.active
companies = []
import glob 

files = list(filter(os.path.isfile, glob.glob(file_path + "\*"))) 
  
# Sorting file list based on the  
# creation time of the files 
files.sort(key=os.path.getctime) 
  
# Displaying the sorted list 
print(files)
        
# for files in os.listdir(file_path):
#     ind = files.find(".")
#     comp = files[:ind]
#     if(comp!=''):
#         present.append(int(comp))

        
# for rows in range(3,1126):
#     if(ws.cell(row = rows, column = 2).value!=None):
#         companies.append(ws.cell(row = rows, column = 2).value)

# for i in range(len(present)):
#     for j in range(len(companies)):
#         if(present[i]== companies[j]):
#             companies[j]=0
# remaining = []
# for i in range(len(companies)):
#     if(companies[i])!=0:
#         remaining.append(companies[i])
# print(remaining)
    
    
    
       
    

    
