import csv
import pandas as pd
import numpy as np
import openpyxl
from keras.models import Sequential
from keras.layers import LSTM
from keras.layers import Dense
#import sys
import pympler
from pympler import asizeof
import timeit
import memory_profiler
open_workbook = '/Users/umang/umang/3-1/random/stock_market/eligible.xlsx'
workbook = openpyxl.load_workbook(open_workbook)
ws = workbook.active
companies = []


class stocks():
    
    def __init__(self, id_number):
        self.id_number = id_number
        self.name = self.find_name()
        self.data_array = self.csv_to_array()
        
        self.train_array = self.data_array[:round(0.8*len(self.data_array))]
        self.test_array = self.data_array[round(0.8*len(self.data_array)):]
        #self.data_csv = pd.read_csv(f"new2/{self.id_number}.csv")
        #self.output, self.mse = self.lstm_model()
        #self.slope_5 = self.last_5_slopes()
        #self.profit = self.calc_profit()
        #self.final_score = self.final_score_fun()

    last_5_slopes = 0
    
        
        
    def csv_to_array(self):
        data_array = []
        csv_in = pd.read_csv(f"new2/{self.id_number}.csv")
        data = csv_in["Close Price"]
        for i in range(len(data)):
            data_array.append(data[i]) 
        
        return data_array
    
    def find_name(self):
        for rows in range(1,451):
            if(ws.cell(row = rows, column = 2).value==self.id_number):
                name = ws.cell(row = rows, column = 3).value
                break
        return name
    
    def lstm_model(self):
        raw_seq = self.train_array
        n_steps = 5
        
        def split_sequence(sequence, n_steps):
            X, y = list(), list()
            for i in range(len(sequence)):
                # find the end of this pattern
                end_ix = i + n_steps
                # check if we are beyond the sequence
                if end_ix > len(sequence)-1:
                    break
                # gather input and output parts of the pattern
                seq_x, seq_y = sequence[i:end_ix], sequence[end_ix]
                X.append(seq_x)
                y.append(seq_y)
            return np.array(X), np.array(y)

               
        X, y = split_sequence(raw_seq, n_steps)
                
        n_features = 1
        X = X.reshape((X.shape[0], X.shape[1], n_features))
                # define model
        model = Sequential()
        model.add(LSTM(50, activation='relu', input_shape=(n_steps, n_features)))
        model.add(Dense(1))
        model.compile(optimizer='adam', loss='mse')
                # fit model
        model.fit(X, y, epochs=200, verbose=0)
                # demonstrate prediction
        
        x_input, _ = split_sequence(self.test_array, n_steps)
        mse = 0
        for i in range(len(self.test_array)-6):
           

        
            x_input_ = x_input[i].reshape((1, n_steps, n_features))

    
            Y_out = model.predict(x_input_, verbose=0)   
            X_out = (self.test_array[6+i])
            
            mse = mse + (((Y_out-X_out)/X_out))**2
        

        return Y_out, mse
        
     
    def last_5_slopes(self):
        length = len(self.data_array)
        slopes = 0
        
        for i in range(1,6):
            slope = (self.data_array[length-i-1]-self.data_array[length-i-2])/self.data_array[length-i-1]
            slopes = slopes + slope*(i-0.5)/15
        return slopes
    def percentage_change(self):
        
        change = self.data_array[len(self.data_array)-1] - self.data_array[0]
        per_change = change/(self.data_array[0])
        return per_change
    # def calc_profit(self):
    #     Y_out, _ = self.lstm_model()
    #     profit = (Y_out - self.data_array[len(self.data_array)-1])/self.data_array[len(self.data_array)-1]
    #     return profit
    def dist_from_av(self):
        sum_of = 0
        for i in range(1,6):
            sum_of = sum_of + self.data_array[len(self.data_array)-i]
        av = sum_of/5
        return av

    def final_score_fun(self):
        _, mse = self.lstm_model()
        profit = (_ - self.data_array[len(self.data_array)-1])/self.data_array[len(self.data_array)-1]
        dist_av_per = (self.dist_from_av()-self.data_array[len(self.data_array)-1])/self.data_array[len(self.data_array)-1]     
        per_change = self.percentage_change()
        #self.calc_profit()
        last_5 = self.last_5_slopes()
        if((profit<0) and (last_5<0)):
            final_score = (profit*last_5*-1)/(mse)
        else:
            final_score = (profit*last_5)/(mse)
        return final_score



def check_empty_row():
    i = 1
    while (ws.cell(row = i, column = 2).value!=None):
        i=i+1
        print(i)
    return i    
names = []

def gather_companies():
    count = 0
    for rows in range(3,451):
        if(ws.cell(row = rows, column = 2).value!=None):
            a = stocks(ws.cell(row = rows, column = 2).value).final_score_fun()
            companies.append(a)
            print("company_appended")
            b = stocks(ws.cell(row = rows, column = 2).value).name
            names.append(b)
            count = count +1

            print(b)
            print(a)
            print(count)
worst_companies = [0,0,0,0,0]
worst_company_name = [0,0,0,0,0]
     
best_companies = [0,0,0,0,0]
best_company_name = [0,0,0,0,0]
def best_comp():
    for i in range(len(companies)):
            if((companies[i])>min(best_companies)):
               idx = best_companies.index(min(best_companies))
               best_companies[idx]=companies[i]
               best_company_name[idx]= names[i]

    return best_companies
def worst_comp():
    for i in range(len(companies)):
            if((companies[i])<max(worst_companies)):
               idx = worst_companies.index(max(worst_companies))
               worst_companies[idx]=companies[i]
               worst_company_name[idx]= names[i]

    return worst_companies
#print(asizeof.asizeof(stocks(500325)))

# a = stocks(500325)
# print(a.final_score)
# print("got company")
gather_companies()
best_comp()
worst_comp()
print(best_companies)
print(best_company_name)
#print(names)
print(worst_companies)
print(worst_company_name)
    


        

        
        


